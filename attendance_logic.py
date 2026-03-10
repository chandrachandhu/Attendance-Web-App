import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

def process_attendance(input_file, output_file):

    sheet1_name = "Sheet1"
    sheet2_name = "Sheet2"
    sheet3_name = "Sheet3"
    sheet4_name = "Sheet4"

    subject_columns = ['D', 'F', 'H', 'J', 'L', 'N', 'P', 'R']
    percentage_offset = 1
    section_headers = [5, 87, 167, 247]

    subject_map = {
        'E': 'D',
        'G': 'E',
        'I': 'F',
        'K': 'G',
        'M': 'H',
        'O': 'I',
        'Q': 'J',
        'S': 'K',
    }

    sections_sheet1 = [
        (6, 82),
        (88, 162),
        (168, 242),
        (248, 322)
    ]

    skip_ranges = [
        range(83, 87),
        range(163, 167),
        range(243, 247)
    ]

    sheet2_start_row = 16

    sections_info = [
        (6, 82, 5, 12),
        (88, 162, 87, 89),
        (168, 242, 167, 164),
        (248, 322, 247, 239)
    ]

    def get_writable_cell(ws, row, col_letter):
        cell = ws[f"{col_letter}{row}"]
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(merged_range.min_row, merged_range.min_col)
        return ws.cell(row, column_index_from_string(col_letter))

    wb = openpyxl.load_workbook(input_file)

    sheet1 = wb[sheet1_name]
    sheet2 = wb[sheet2_name]
    sheet3 = wb[sheet3_name]
    sheet4 = wb[sheet4_name]

    # -----------------------
    # STEP 1
    # -----------------------

    for header_row in section_headers:

        data_row = header_row

        while sheet1[f"A{data_row}"].value not in (None, "", " "):

            total_attended = 0
            total_possible = 0

            for col in subject_columns:

                held_val = sheet1[f"{col}{header_row}"].value
                attended_val = sheet1[f"{col}{data_row}"].value

                perc_col = get_column_letter(column_index_from_string(col) + percentage_offset)

                perc_cell = get_writable_cell(sheet1, data_row, perc_col)

                try:

                    held = float(held_val or 0)
                    attended = float(attended_val or 0)

                    if data_row == header_row:
                        perc_cell.value = 100.0 if held > 0 else 0.0
                    else:
                        percent = round((attended / held) * 100, 2) if held > 0 else 0.0
                        perc_cell.value = percent

                    total_attended += attended
                    total_possible += held

                except:
                    continue

            total_attended_col = 'T'
            total_percentage_col = 'U'

            total_cell = get_writable_cell(sheet1, data_row, total_attended_col)
            percent_cell = get_writable_cell(sheet1, data_row, total_percentage_col)

            if data_row == header_row:

                total_cell.value = int(total_possible)

                percent_cell.value = 100.0 if total_possible > 0 else 0.0

            else:

                total_cell.value = int(total_attended)

                percent_cell.value = round((total_attended / total_possible) * 100, 2) if total_possible > 0 else 0.0

            data_row += 1

    # -----------------------
    # STEP 2
    # -----------------------

    for s1_col, s2_col in subject_map.items():

        row2 = sheet2_start_row

        for start_row, end_row in sections_sheet1:

            for row1 in range(start_row, end_row + 1):

                if any(row1 in skip_range for skip_range in skip_ranges):
                    continue

                value = sheet1[f"{s1_col}{row1}"].value

                try:
                    value = float(value or 0)
                except:
                    value = 0.0

                sheet2[f"{s2_col}{row2}"] = value

                row2 += 1

    # -----------------------
    # STEP 3
    # -----------------------

    for start_s1, end_s1, header_s1, start_s3 in sections_info:

        total_classes = sheet1[f"T{header_s1}"].value or 0

        for r in range(start_s3, start_s3 + (end_s1 - start_s1) + 1):
            get_writable_cell(sheet3, r, "G").value = total_classes

        s3_row = start_s3

        for r in range(start_s1, end_s1 + 1):

            attended = sheet1[f"T{r}"].value or 0

            get_writable_cell(sheet3, s3_row, "H").value = attended

            s3_row += 1

        s3_row = start_s3

        for r in range(start_s1, end_s1 + 1):

            percent = sheet1[f"U{r}"].value or 0

            get_writable_cell(sheet3, s3_row, "I").value = percent

            s3_row += 1

    # -----------------------
    # STEP 4
    # -----------------------

    start_row_monthly = 12

    num_students = 0

    while True:

        cell_val = sheet3[f"A{start_row_monthly + num_students}"].value

        if cell_val is None or str(cell_val).strip() == "":
            break

        num_students += 1

    for i in range(num_students):

        row = start_row_monthly + i

        june_held = sheet3[f"G{row}"].value or 0
        june_attended = sheet3[f"H{row}"].value or 0

        may_held = sheet4[f"G{row}"].value or 0
        may_attended = sheet4[f"H{row}"].value or 0

        monthly_held = float(june_held) - float(may_held)
        monthly_attended = float(june_attended) - float(may_attended)

        if monthly_held > 0:
            monthly_percent = round((monthly_attended / monthly_held) * 100, 2)
        else:
            monthly_percent = 0.0

        sheet3[f"D{row}"] = monthly_held
        sheet3[f"E{row}"] = monthly_attended
        sheet3[f"F{row}"] = monthly_percent

    # -----------------------
    # STEP 5
    # -----------------------

    red_fill = PatternFill(start_color="FF7D7D", end_color="FF7D7D", fill_type="solid")

    for i in range(num_students):

        row = start_row_monthly + i

        row_values = [sheet3.cell(row=row, column=col).value for col in range(1, sheet3.max_column + 1)]

        condition1 = any((isinstance(v, (int, float)) and v <= 0) for v in row_values if v is not None)

        d_val = sheet3[f"D{row}"].value or 0
        e_val = sheet3[f"E{row}"].value or 0

        condition2 = float(e_val) > float(d_val)

        if condition1 or condition2:

            for col in range(1, sheet3.max_column + 1):
                sheet3.cell(row=row, column=col).fill = red_fill

    wb.save(output_file)